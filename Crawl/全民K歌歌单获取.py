import re
import requests
from bs4 import BeautifulSoup
import openpyxl  # 操作xlsx格式的Excel
import json
# ugc_total_count代表歌曲总数量
# has_more代表歌曲是否显示完毕，是1否0
num = 0  # 歌曲总数量
p = 0  # 歌曲最少页数
q = 0  # 最后一页的歌曲数

kv = {
    "User-Agent":
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"
}


# 获取share_uid
def getShare_UID(url):
    uid = url.split('=')[-1]
    # print("uid={}".format(uid))
    return uid


# 获取网页
def getHTML(url):
    try:
        r = requests.get(url, headers=kv, timeout=30)
        r.raise_for_status()
        # print(r.url)
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ""


# 解析json数据获取歌曲名称信息
def getMusicInfo(json_data, namelis, idlis, count):
    s = json.loads(json_data)  # 将json数据转换为python对象(字典)
    # print(s)
    # print(type(s))
    # 获取歌曲名称
    # num = s["data"]["ugc_total_count"]
    # print(num)
    for i in range(count):
        m_name = s["data"]["ugclist"][i]["title"]
        m_shareid = s["data"]["ugclist"][i]["shareid"]
        namelis.append(m_name)
        idlis.append(m_shareid)
        # print(m_name)
        # print(m_shareid)


# 打印歌曲信息
def printMusicInfo(namelis, idlis, music_url):
    form = "{:^3}\t{:^15}\t{:^20}"
    print("{}的演唱歌曲共有{}首，信息如下：".format(nickname, num))
    print(form.format("序号", "歌曲名称", "下载链接"))
    for i in range(len(namelis)):
        down_url = music_url + idlis[i]
        print(form.format(i + 1, namelis[i], down_url))


# 将歌曲信息保存到Excel文件
def saveXlsx(namelis, idlis, music_url):
    try:
        wb = openpyxl.Workbook()  # 创建一个工作簿
        ws = wb.active  # 在工作簿中获取活跃的工作表
        ws.title = "歌曲信息"  # 设置工作表的名称
        ws['A1'] = '序号'
        ws['B1'] = '歌曲名称'
        ws['C1'] = '下载链接'
        for i in range(len(namelis)):
            ws.cell(row=i + 2, column=1, value=i + 1)  # 行数从2开始
            ws.cell(row=i + 2, column=2, value=namelis[i])  #歌曲名称
            ws.cell(row=i + 2, column=3, value=music_url + idlis[i])  # 歌曲链接
        wb.save('D:/{}的歌曲信息.xlsx'.format(nickname_cn))
        print("歌曲信息在D盘根目录保存完毕！")
    except:
        print("出现异常，保存失败！")


if __name__ == "__main__":
    namelis = []
    idlis = []
    home_url=input("请输入作者的主页链接：")
    # home_url = "http://kg.qq.com/node/personal?uid=619f9f8c2429318d35"
    music_url = "https://node.kg.qq.com/cgi/fcgi-bin/fcg_get_play_url?shareid="
    kg_url = "https://node.kg.qq.com/cgi/fcgi-bin/kg_ugc_get_homepage?jsonpCallback=callback_0&inChFarset=GB2312&outCharset=utf-8&format=&g_tk=5381&g_tk_openkey=719182536&nocache=0.8706501019187272&share_uid="
    kg2_url = kg_url + getShare_UID(home_url) + "&type=get_uinfo&start=1&num=8"
    # print(kg2_url)
    home_json = getHTML(kg2_url)[11:-1]  # json数据在callback_0()里面，需要单独提取出来
    # print(home_json)
    # with open("D:\VscodePy\pytest\KG.txt", "r", encoding='utf-8') as f:
    #     home_json = f.read()
    #     f.close()
    s = json.loads(home_json)  # 将json数据转换为python对象(字典)
    num = s["data"]["ugc_total_count"]  # 获取歌曲总数量
    nickname = s["data"]["nickname"]  # 获取昵称
    match = re.search(r'[\u4e00-\u9fa5\d]*', nickname)  # 文件名称不能有特殊字符
    nickname_cn = match.group(0)
    count = 8
    index = 1
    q = num // 8
    p = num % 8
    if q < 1:  # 说明p!=0
        x = 1
        # count = p
    elif p == 0:  # 说明歌曲数量是8的倍数
        x = q
    else:  # 多一页刚好爬完
        x = q + 1
    for i in range(x):
        if i == x - 1:
            if p != 0:
                count = p
        kg2_url = kg_url + getShare_UID(
            home_url) + "&type=get_uinfo&start=" + str(index) + "&num=8"
        home_json = getHTML(kg2_url)[11:-1]
        getMusicInfo(home_json, namelis, idlis, count)
        index += 1
    printMusicInfo(namelis, idlis, music_url)
    saveXlsx(namelis, idlis, music_url)
